"""
Analytics Celery Tasks

Asynchronous tasks for analytics report generation, data export,
and dashboard metric calculation.
"""

import os
import json
import logging
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from celery import shared_task
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.conf import settings
from django.utils import timezone
from django.template.loader import render_to_string
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import AnalyticsReport, AnalyticsMetric
from .services import CrossModuleAnalyticsService, AnalyticsReportGenerator
from risk.analytics import RiskAnalyticsService

logger = logging.getLogger(__name__)


class DecimalEncoder(json.JSONEncoder):
    """JSON encoder for Decimal values."""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


@shared_task(bind=True, max_retries=3)
def generate_analytics_report(self, report_id):
    """
    Generate an analytics report in the requested format.

    Args:
        report_id: UUID of the AnalyticsReport to generate
    """
    try:
        report = AnalyticsReport.objects.get(id=report_id)
        report.mark_started()

        logger.info(f"Starting generation of {report.report_type} report in {report.export_format} format")

        # Get the appropriate data based on report type
        data = _get_report_data(report.report_type, report.filters)

        # Generate the report file
        if report.export_format == 'pdf':
            file_path = _generate_pdf_report(report, data)
        elif report.export_format == 'excel':
            file_path = _generate_excel_report(report, data)
        elif report.export_format == 'csv':
            file_path = _generate_csv_export(report, data)
        elif report.export_format == 'json':
            file_path = _generate_json_export(report, data)
        else:
            raise ValueError(f"Unsupported export format: {report.export_format}")

        # Calculate file size
        file_size = default_storage.size(file_path)
        data_points = _count_data_points(data)

        # Mark as completed
        report.mark_completed(file_path, file_size, data_points)

        logger.info(f"Successfully generated report {report_id} at {file_path}")
        return {
            'status': 'completed',
            'file_path': file_path,
            'file_size': file_size,
            'data_points': data_points
        }

    except AnalyticsReport.DoesNotExist:
        logger.error(f"Analytics report {report_id} not found")
        return {'status': 'failed', 'error': 'Report not found'}

    except Exception as e:
        logger.error(f"Error generating report {report_id}: {str(e)}")

        try:
            report = AnalyticsReport.objects.get(id=report_id)
            report.mark_failed(str(e))
        except AnalyticsReport.DoesNotExist:
            pass

        # Retry on failure
        if self.request.retries < self.max_retries:
            logger.info(f"Retrying report generation for {report_id}")
            raise self.retry(countdown=60, exc=e)

        return {'status': 'failed', 'error': str(e)}


@shared_task
def cleanup_expired_reports():
    """
    Clean up expired analytics reports and their associated files.
    """
    try:
        expired_reports = AnalyticsReport.objects.filter(
            expires_at__lt=timezone.now(),
            status='completed'
        )

        cleaned_count = 0
        for report in expired_reports:
            try:
                # Delete the file if it exists
                if report.file_path and default_storage.exists(report.file_path):
                    default_storage.delete(report.file_path)

                # Delete the report record
                report.delete()
                cleaned_count += 1

            except Exception as e:
                logger.error(f"Error cleaning up report {report.id}: {str(e)}")

        logger.info(f"Cleaned up {cleaned_count} expired analytics reports")
        return {'cleaned_count': cleaned_count}

    except Exception as e:
        logger.error(f"Error in cleanup_expired_reports: {str(e)}")
        return {'error': str(e)}


@shared_task
def cache_analytics_metrics():
    """
    Pre-calculate and cache analytics metrics for improved dashboard performance.
    """
    try:
        cached_count = 0
        calculation_start = timezone.now()

        # Cache executive dashboard metrics
        try:
            executive_data = CrossModuleAnalyticsService.get_executive_dashboard_data()
            _cache_metric('integrated', 'Executive Dashboard', 'executive_summary', executive_data)
            cached_count += 1
        except Exception as e:
            logger.error(f"Error caching executive metrics: {str(e)}")

        # Cache compliance metrics
        try:
            compliance_data = CrossModuleAnalyticsService.get_compliance_dashboard_data()
            _cache_metric('compliance', 'Compliance Overview', 'compliance_summary', compliance_data)
            cached_count += 1
        except Exception as e:
            logger.error(f"Error caching compliance metrics: {str(e)}")

        # Cache risk analytics
        try:
            risk_data = RiskAnalyticsService.get_executive_risk_summary()
            _cache_metric('risk', 'Risk Executive Summary', 'risk_executive', risk_data)
            cached_count += 1
        except Exception as e:
            logger.error(f"Error caching risk metrics: {str(e)}")

        # Cache vendor risk data
        try:
            vendor_data = CrossModuleAnalyticsService.get_vendor_risk_dashboard_data()
            _cache_metric('vendor', 'Vendor Risk Summary', 'vendor_risk_summary', vendor_data)
            cached_count += 1
        except Exception as e:
            logger.error(f"Error caching vendor metrics: {str(e)}")

        # Cache policy metrics
        try:
            policy_data = CrossModuleAnalyticsService.get_policy_management_dashboard_data()
            _cache_metric('policy', 'Policy Management Summary', 'policy_summary', policy_data)
            cached_count += 1
        except Exception as e:
            logger.error(f"Error caching policy metrics: {str(e)}")

        # Cache training metrics
        try:
            training_data = CrossModuleAnalyticsService.get_training_effectiveness_dashboard_data()
            _cache_metric('training', 'Training Effectiveness Summary', 'training_summary', training_data)
            cached_count += 1
        except Exception as e:
            logger.error(f"Error caching training metrics: {str(e)}")

        calculation_time = (timezone.now() - calculation_start).total_seconds()

        logger.info(f"Successfully cached {cached_count} analytics metrics in {calculation_time:.2f}s")
        return {
            'cached_count': cached_count,
            'calculation_time_seconds': calculation_time
        }

    except Exception as e:
        logger.error(f"Error in cache_analytics_metrics: {str(e)}")
        return {'error': str(e)}


def _get_report_data(report_type, filters):
    """Get analytics data based on report type."""
    if report_type == 'executive':
        return AnalyticsReportGenerator.generate_executive_report_data()
    elif report_type == 'compliance':
        return CrossModuleAnalyticsService.get_compliance_dashboard_data()
    elif report_type == 'risk':
        return RiskAnalyticsService.get_risk_overview_stats()
    elif report_type == 'vendor':
        return CrossModuleAnalyticsService.get_vendor_risk_dashboard_data()
    elif report_type == 'policy':
        return CrossModuleAnalyticsService.get_policy_management_dashboard_data()
    elif report_type == 'training':
        return CrossModuleAnalyticsService.get_training_effectiveness_dashboard_data()
    elif report_type == 'operational':
        return AnalyticsReportGenerator.generate_operational_dashboard_data()
    elif report_type == 'integrated':
        return CrossModuleAnalyticsService.get_integrated_risk_posture()
    else:
        raise ValueError(f"Unknown report type: {report_type}")


def _generate_pdf_report(report, data):
    """Generate a PDF report from analytics data."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=18)

    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        textColor=colors.HexColor('#2F6FED')
    )

    # Build the PDF content
    story = []

    # Title page
    story.append(Paragraph(report.title, title_style))
    story.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    story.append(Spacer(1, 20))

    if report.description:
        story.append(Paragraph("Report Description", styles['Heading2']))
        story.append(Paragraph(report.description, styles['Normal']))
        story.append(Spacer(1, 20))

    # Add content based on report type
    if report.report_type == 'executive':
        _add_executive_pdf_content(story, data, styles)
    elif report.report_type == 'compliance':
        _add_compliance_pdf_content(story, data, styles)
    elif report.report_type == 'risk':
        _add_risk_pdf_content(story, data, styles)
    # Add more report types as needed

    # Build PDF
    doc.build(story)

    # Save to storage
    buffer.seek(0)
    file_name = f"analytics_reports/{report.id}/{report.report_type}_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    file_path = default_storage.save(file_name, ContentFile(buffer.getvalue()))

    return file_path


def _generate_excel_report(report, data):
    """Generate an Excel report from analytics data."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create summary sheet
    summary_sheet = wb.create_sheet("Executive Summary")
    _add_excel_summary(summary_sheet, report, data)

    # Add data sheets based on report type
    if report.report_type == 'executive':
        _add_executive_excel_sheets(wb, data)
    elif report.report_type == 'compliance':
        _add_compliance_excel_sheets(wb, data)
    elif report.report_type == 'risk':
        _add_risk_excel_sheets(wb, data)

    # Save to storage
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    file_name = f"analytics_reports/{report.id}/{report.report_type}_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = default_storage.save(file_name, ContentFile(buffer.getvalue()))

    return file_path


def _generate_csv_export(report, data):
    """Generate CSV export from analytics data."""
    # Convert data to DataFrame
    df = pd.DataFrame(_flatten_data_for_csv(data))

    # Save to storage
    buffer = BytesIO()
    df.to_csv(buffer, index=False)
    buffer.seek(0)

    file_name = f"analytics_reports/{report.id}/{report.report_type}_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    file_path = default_storage.save(file_name, ContentFile(buffer.getvalue()))

    return file_path


def _generate_json_export(report, data):
    """Generate JSON export from analytics data."""
    # Add metadata
    export_data = {
        'report_metadata': {
            'type': report.report_type,
            'generated_at': timezone.now().isoformat(),
            'title': report.title,
            'description': report.description,
        },
        'data': data
    }

    # Convert to JSON
    json_data = json.dumps(export_data, indent=2, cls=DecimalEncoder)

    # Save to storage
    file_name = f"analytics_reports/{report.id}/{report.report_type}_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
    file_path = default_storage.save(file_name, ContentFile(json_data.encode('utf-8')))

    return file_path


def _cache_metric(category, name, key, value):
    """Cache a calculated metric."""
    AnalyticsMetric.objects.update_or_create(
        category=category,
        metric_key=key,
        defaults={
            'metric_name': name,
            'value': value,
            'calculation_date': timezone.now(),
            'data_source': 'analytics_service',
            'cache_expires_at': timezone.now() + timedelta(hours=1),
            'is_cached': True
        }
    )


def _count_data_points(data):
    """Count the number of data points in the analytics data."""
    if isinstance(data, dict):
        return len(json.dumps(data, cls=DecimalEncoder))
    elif isinstance(data, list):
        return len(data)
    else:
        return 1


def _add_executive_pdf_content(story, data, styles):
    """Add executive report content to PDF."""
    story.append(Paragraph("Executive Summary", styles['Heading2']))
    story.append(Spacer(1, 12))

    # Add key metrics
    if 'executive_dashboard' in data:
        exec_data = data['executive_dashboard']

        # Risk summary
        if 'risk_summary' in exec_data:
            risk_data = exec_data['risk_summary']
            story.append(Paragraph("Risk Management Overview", styles['Heading3']))
            story.append(Paragraph(f"Total Active Risks: {risk_data.get('active_risks', 0)}", styles['Normal']))
            story.append(Paragraph(f"Critical/High Risks: {risk_data.get('critical_high_risks', 0)}", styles['Normal']))
            story.append(Spacer(1, 12))


def _add_compliance_pdf_content(story, data, styles):
    """Add compliance report content to PDF."""
    story.append(Paragraph("Compliance Overview", styles['Heading2']))
    # Add compliance-specific content


def _add_risk_pdf_content(story, data, styles):
    """Add risk report content to PDF."""
    story.append(Paragraph("Risk Assessment", styles['Heading2']))
    # Add risk-specific content


def _add_excel_summary(sheet, report, data):
    """Add summary information to Excel sheet."""
    # Header
    sheet['A1'] = report.title
    sheet['A1'].font = Font(size=16, bold=True)

    sheet['A3'] = "Generated:"
    sheet['B3'] = timezone.now().strftime('%B %d, %Y at %I:%M %p')

    if report.description:
        sheet['A5'] = "Description:"
        sheet['B5'] = report.description


def _add_executive_excel_sheets(wb, data):
    """Add executive data sheets to Excel workbook."""
    if 'executive_dashboard' in data:
        exec_sheet = wb.create_sheet("Executive Metrics")
        # Add executive metrics data


def _add_compliance_excel_sheets(wb, data):
    """Add compliance data sheets to Excel workbook."""
    comp_sheet = wb.create_sheet("Compliance Data")
    # Add compliance data


def _add_risk_excel_sheets(wb, data):
    """Add risk data sheets to Excel workbook."""
    risk_sheet = wb.create_sheet("Risk Data")
    # Add risk data


def _flatten_data_for_csv(data):
    """Flatten nested data structure for CSV export."""
    flattened = []

    def flatten_dict(d, parent_key='', sep='_'):
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list) and v and isinstance(v[0], dict):
                # Handle list of dictionaries
                for i, item in enumerate(v):
                    items.extend(flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)

    if isinstance(data, dict):
        flattened.append(flatten_dict(data))
    elif isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                flattened.append(flatten_dict(item))

    return flattened