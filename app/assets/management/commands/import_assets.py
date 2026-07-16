import hashlib
import ipaddress
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import load_workbook

from assets.models import Asset
from core.audit import log_audit_event

User = get_user_model()


class Command(BaseCommand):
    help = 'Import information assets from the Axim asset register spreadsheet'

    header_aliases = {
        'computer name': 'name',
        'asset name': 'name',
        'assetname': 'name',
        'asset type': 'asset_type',
        'asset typename': 'asset_type',
        'assettypename': 'asset_type',
        'domain': 'domain',
        'ip address': 'ip_address',
        'ipaddress': 'ip_address',
        'last successful scan': 'last_seen_at',
        'lastseen': 'last_seen_at',
        'username': 'custodian',
        'location': 'location',
        'physical memory mb': 'physical_memory_mb',
        'manufacturer': 'manufacturer',
        'monitor manufacturer': 'manufacturer',
        'monitormanufacturer': 'manufacturer',
        'device model': 'model',
        'monitor model': 'model',
        'monitormodel': 'model',
        'version': 'version',
        'operating system': 'operating_system',
        'system type': 'system_type',
        'serial number': 'serial_number',
        'serialnumber': 'serial_number',
        'service pack': 'service_pack',
        'mac address': 'mac_address',
        'mac': 'mac_address',
        'criticality': 'criticality',
        'owner': 'owner_name',
        'system description': 'description',
        'description': 'description',
        'comment': 'comments',
        'comments': 'comments',
    }

    sheet_asset_types = {
        'server': 'server',
        'workstation': 'workstation',
        'desktop': 'workstation',
        'laptop': 'workstation',
        'monitor': 'monitor',
        'mobile': 'mobile_device',
        'printer': 'printer',
        'infrastructure': 'infrastructure',
    }

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the asset register .xlsx file')
        parser.add_argument(
            '--user',
            type=str,
            default='system',
            help='Username to attribute imported assets to',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Show what would be imported without writing assets',
        )

    def handle(self, *args, **options):
        file_path = Path(options['file_path'])
        if not file_path.exists():
            raise CommandError(f'File not found: {file_path}')
        if file_path.suffix.lower() != '.xlsx':
            raise CommandError('Asset imports require a .xlsx file.')

        entries = self.collect_entries(file_path)

        if options['dry_run']:
            self.show_dry_run_summary(entries)
            return

        user = self.get_import_user(options['user'])
        with transaction.atomic():
            imported, updated = self.import_entries(entries, file_path, user)
            self.record_import_audit_event(file_path, entries, user, imported, updated)

        self.stdout.write(
            self.style.SUCCESS(f'Imported {imported} assets; updated {updated}.')
        )

    def get_import_user(self, username):
        if username != 'system':
            try:
                return User.objects.get(username=username)
            except User.DoesNotExist:
                raise CommandError(f'User not found: {username}')

        user, created = User.objects.get_or_create(
            username='system',
            defaults={'email': 'system@localhost', 'is_active': False},
        )
        if created:
            user.set_unusable_password()
            user.save(update_fields=['password'])
        return user

    def collect_entries(self, file_path, source_label=None):
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        entries = []
        skipped_count = 0
        sheet_counts = {}
        recognised_headers = 0
        source_label = source_label or str(file_path)

        for worksheet in workbook.worksheets:
            header_row, header_map = self.find_header_row(worksheet)
            if not header_map:
                continue
            recognised_headers += 1

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                row_data = self.extract_row(row, header_map, worksheet.title)
                if not row_data:
                    skipped_count += 1
                    continue

                row_data['source_path'] = source_label
                row_data['source_sheet'] = worksheet.title
                row_data['source_row'] = row_number
                row_data['source_checksum'] = self.calculate_row_checksum(row_data)
                entries.append(row_data)
                sheet_counts[worksheet.title] = sheet_counts.get(worksheet.title, 0) + 1

        self.skipped_count = skipped_count
        self.sheet_counts = sheet_counts
        if recognised_headers == 0:
            raise CommandError('No recognised asset header row found.')
        return entries

    def find_header_row(self, worksheet):
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            header_map = {}
            for index, value in enumerate(row):
                normalised = self.normalise_header(value)
                canonical = self.header_aliases.get(normalised)
                if canonical:
                    header_map[canonical] = index
            if 'name' in header_map:
                return row_number, header_map
        return None, {}

    def extract_row(self, row, header_map, sheet_name):
        name = self.clean_text(self.get_cell(row, header_map, 'name'))
        if not name:
            return {}

        serial_number = self.clean_text(self.get_cell(row, header_map, 'serial_number'))
        mac_address = self.clean_text(self.get_cell(row, header_map, 'mac_address'))
        ip_address = self.clean_ip_address(self.get_cell(row, header_map, 'ip_address'))
        asset_id = self.build_asset_id(name, serial_number, mac_address, ip_address)
        criticality = self.normalise_choice(
            self.get_cell(row, header_map, 'criticality'),
            {'low', 'medium', 'high', 'critical'},
            'medium',
        )
        owner_name = self.clean_text(self.get_cell(row, header_map, 'owner_name'))

        metadata = {
            'sheet_asset_type': self.infer_asset_type(sheet_name),
            'system_type': self.clean_text(self.get_cell(row, header_map, 'system_type')),
            'service_pack': self.clean_text(self.get_cell(row, header_map, 'service_pack')),
            'physical_memory_mb': self.clean_text(self.get_cell(row, header_map, 'physical_memory_mb')),
            'comments': self.clean_text(self.get_cell(row, header_map, 'comments')),
        }

        return {
            'asset_id': asset_id,
            'name': name[:255],
            'asset_type': self.normalise_asset_type(
                self.get_cell(row, header_map, 'asset_type'),
                sheet_name,
            ),
            'description': self.clean_text(self.get_cell(row, header_map, 'description')),
            'classification': 'internal',
            'criticality': criticality,
            'lifecycle_status': 'active',
            'owner': self.resolve_owner(owner_name),
            'owner_name': owner_name,
            'custodian': self.clean_text(self.get_cell(row, header_map, 'custodian')),
            'location': self.clean_text(self.get_cell(row, header_map, 'location')),
            'domain': self.clean_text(self.get_cell(row, header_map, 'domain')),
            'ip_address': ip_address,
            'mac_address': mac_address[:64],
            'serial_number': serial_number[:128],
            'manufacturer': self.clean_text(self.get_cell(row, header_map, 'manufacturer'))[:255],
            'model': self.clean_text(self.get_cell(row, header_map, 'model'))[:255],
            'operating_system': self.clean_text(self.get_cell(row, header_map, 'operating_system'))[:255],
            'version': self.clean_text(self.get_cell(row, header_map, 'version'))[:100],
            'last_seen_at': self.parse_datetime_value(self.get_cell(row, header_map, 'last_seen_at')),
            'metadata': {key: value for key, value in metadata.items() if value},
        }

    def import_entries(self, entries, file_path, user):
        imported = 0
        updated = 0

        for entry in entries:
            values = {
                **entry,
                'created_by': user,
            }
            asset, created = Asset.objects.update_or_create(
                asset_id=entry['asset_id'],
                defaults=values,
            )
            if created:
                imported += 1
            else:
                updated += 1

        return imported, updated

    def record_import_audit_event(self, file_path, entries, user, imported, updated):
        log_audit_event(
            actor=user,
            event='ASSET_REGISTER_IMPORTED',
            object_type='assets.Asset',
            source={'type': 'import', 'reference': str(file_path)},
            details={
                'source_file': str(file_path),
                'imported_count': imported,
                'updated_count': updated,
                'skipped_count': self.skipped_count,
                'total_importable': len(entries),
                'sheets': self.sheet_counts,
            },
        )

    def show_dry_run_summary(self, entries):
        self.stdout.write(self.style.WARNING('DRY RUN - No changes will be made'))
        self.stdout.write(f'Importable assets: {len(entries)}')
        self.stdout.write(f'Skipped rows: {self.skipped_count}')
        self.stdout.write(f'Sheets: {self.sheet_counts}')
        for entry in entries[:5]:
            self.stdout.write(
                f'  - {entry["source_sheet"]}: {entry["asset_id"]} {entry["name"]}'
            )

    def get_cell(self, row, header_map, field):
        index = header_map.get(field)
        if index is None or index >= len(row):
            return ''
        return row[index]

    def clean_text(self, value):
        if value is None:
            return ''
        if isinstance(value, Decimal):
            return format(value.normalize(), 'f')
        return re.sub(r'\s+', ' ', str(value).strip())

    def normalise_header(self, value):
        return re.sub(r'[^a-z0-9]+', ' ', str(value or '').strip().lower()).strip()

    def normalise_choice(self, value, valid_choices, default):
        normalised = self.clean_text(value).lower().replace('-', '_').replace(' ', '_')
        return normalised if normalised in valid_choices else default

    def normalise_asset_type(self, value, sheet_name):
        explicit = self.clean_text(value).lower().replace('-', '_').replace(' ', '_')
        if explicit in {choice[0] for choice in Asset.ASSET_TYPES}:
            return explicit
        return self.infer_asset_type(sheet_name)

    def infer_asset_type(self, sheet_name):
        normalised = sheet_name.lower()
        for token, asset_type in self.sheet_asset_types.items():
            if token in normalised:
                return asset_type
        return 'other'

    def clean_ip_address(self, value):
        text = self.clean_text(value)
        if not text or text.lower() in {'none', 'n/a'}:
            return None
        try:
            ipaddress.ip_address(text)
        except ValueError:
            return None
        return text

    def resolve_owner(self, owner_name):
        if not owner_name:
            return None
        return User.objects.filter(
            Q(username__iexact=owner_name) | Q(email__iexact=owner_name)
        ).first()

    def build_asset_id(self, name, serial_number, mac_address, ip_address):
        stable_key = serial_number or mac_address or ip_address or name
        return re.sub(r'[^A-Z0-9]+', '-', stable_key.upper()).strip('-')[:80]

    def calculate_row_checksum(self, row_data):
        digest_fields = [
            row_data.get('asset_id', ''),
            row_data.get('name', ''),
            row_data.get('asset_type', ''),
            row_data.get('serial_number', ''),
            row_data.get('mac_address', ''),
            row_data.get('ip_address') or '',
        ]
        return hashlib.sha256('|'.join(digest_fields).encode('utf-8')).hexdigest()

    def parse_datetime_value(self, value):
        if not value:
            return None
        if isinstance(value, datetime):
            if timezone.is_aware(value):
                return value
            return timezone.make_aware(value, timezone.get_current_timezone())

        text = self.clean_text(value)
        parsed_datetime = parse_datetime(text)
        if parsed_datetime:
            if timezone.is_aware(parsed_datetime):
                return parsed_datetime
            return timezone.make_aware(parsed_datetime, timezone.get_current_timezone())

        parsed_date = parse_date(text)
        if parsed_date:
            return timezone.make_aware(
                datetime.combine(parsed_date, datetime.min.time()),
                timezone.get_current_timezone(),
            )
        return None
