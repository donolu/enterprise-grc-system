from django.test import TestCase
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.urls import reverse
from django.utils import timezone
from django_tenants.test.cases import TenantTestCase
from rest_framework.test import APIClient
from rest_framework import status
from unittest.mock import patch, MagicMock
from io import BytesIO
import zipfile

from openpyxl import load_workbook

from catalogs.models import Framework, Clause, Control, ControlAssessment, ControlEvidence, AssessmentEvidence
from compliance.models import GovernanceArtefact, RegulatoryRequirement
from core.models import AuditEvent, Document
from assets.models import Asset
from risk.models import Risk
from .models import AssessmentReport, TenantDataExport
from .services import AssessmentReportGenerator, TenantDataExportGenerator, get_export_coverage_manifest
from .tasks import cleanup_old_tenant_data_exports

User = get_user_model()


def response_results(payload):
    return payload.get('results', payload) if isinstance(payload, dict) else payload


class ExportTenantAPITestCase(TenantTestCase):
    """Base API test case that routes requests through a tenant host."""

    def setUp(self):
        super().setUp()
        self.client = APIClient()
        self.client.defaults['HTTP_HOST'] = self.domain.domain


class AssessmentReportModelTest(TestCase):
    """Test AssessmentReport model."""
    
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.framework = Framework.objects.create(
            name='Test Framework',
            short_name='TEST',
            version='1.0',
            effective_date=timezone.now().date()
        )
    
    def test_assessment_report_creation(self):
        """Test creating an AssessmentReport."""
        report = AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Test Summary Report',
            description='Test description',
            framework=self.framework,
            requested_by=self.user
        )
        
        self.assertEqual(report.report_type, 'assessment_summary')
        self.assertEqual(report.title, 'Test Summary Report')
        self.assertEqual(report.framework, self.framework)
        self.assertEqual(report.requested_by, self.user)
        self.assertEqual(report.status, 'pending')
        self.assertTrue(report.include_evidence_summary)
    
    def test_string_representation(self):
        """Test the string representation of AssessmentReport."""
        report = AssessmentReport.objects.create(
            report_type='detailed_assessment',
            title='Detailed Report',
            requested_by=self.user
        )
        
        expected = 'Detailed Assessment Report - Detailed Report'
        self.assertEqual(str(report), expected)


class AssessmentReportAPITest(ExportTenantAPITestCase):
    """Test AssessmentReport API endpoints."""
    
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        
        # Create test framework and assessment data
        self.framework = Framework.objects.create(
            name='ISO 27001',
            short_name='ISO27001',
            version='2022',
            effective_date=timezone.now().date(),
            status='active',
        )
        
        self.clause = Clause.objects.create(
            framework=self.framework,
            clause_id='A.5.1',
            title='Test Clause',
            description='Test clause description'
        )
        
        self.control = Control.objects.create(
            control_id='A.5.1.1',
            name='Test Control',
            description='Test control description',
            control_type='preventive'
        )
        self.control.clauses.add(self.clause)
        
        self.assessment = ControlAssessment.objects.create(
            control=self.control,
            assigned_to=self.user,
            status='complete',
            implementation_status='implemented',
            implementation_approach='Test implementation'
        )
        
        self.client.force_authenticate(user=self.user)
    
    def test_create_assessment_report(self):
        """Test creating an assessment report via API."""
        url = reverse('assessment-reports-list')
        data = {
            'report_type': 'assessment_summary',
            'title': 'Test API Report',
            'description': 'Created via API',
            'framework': self.framework.id,
            'include_evidence_summary': True
        }
        
        response = self.client.post(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(AssessmentReport.objects.count(), 1)
        
        report = AssessmentReport.objects.first()
        self.assertEqual(report.title, 'Test API Report')
        self.assertEqual(report.framework, self.framework)
        self.assertEqual(report.requested_by, self.user)
        event = AuditEvent.objects.get(event='ASSESSMENT_REPORT_REQUESTED')
        self.assertEqual(event.details['actor']['email'], self.user.email)
        self.assertEqual(event.details['object']['type'], 'exports.AssessmentReport')
        self.assertEqual(event.details['new']['title'], 'Test API Report')
    
    def test_list_assessment_reports(self):
        """Test listing assessment reports."""
        # Create test reports
        AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Report 1',
            framework=self.framework,
            requested_by=self.user
        )
        AssessmentReport.objects.create(
            report_type='detailed_assessment',
            title='Report 2',
            requested_by=self.user
        )
        
        url = reverse('assessment-reports-list')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response_results(response.data)), 2)
    
    def test_filter_reports_by_type(self):
        """Test filtering reports by report type."""
        AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Summary Report',
            requested_by=self.user
        )
        AssessmentReport.objects.create(
            report_type='detailed_assessment',
            title='Detailed Report',
            requested_by=self.user
        )
        
        url = reverse('assessment-reports-list')
        response = self.client.get(url, {'report_type': 'assessment_summary'})
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        results = response_results(response.data)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['report_type'], 'assessment_summary')
    
    @patch('exports.tasks.generate_assessment_report_task.delay')
    def test_generate_report(self, mock_task):
        """Test triggering report generation."""
        report = AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Test Report',
            framework=self.framework,
            requested_by=self.user
        )
        
        url = reverse('assessment-reports-generate', kwargs={'pk': report.id})
        response = self.client.post(url)
        
        self.assertEqual(response.status_code, status.HTTP_202_ACCEPTED)
        self.assertIn('Report generation started', response.data['message'])
        
        # Verify task was called
        mock_task.assert_called_once_with(report.id)
        
        # Verify status updated
        report.refresh_from_db()
        self.assertEqual(report.status, 'processing')
        self.assertIsNotNone(report.generation_started_at)
    
    def test_generate_already_processing_report(self):
        """Test generating a report that's already processing."""
        report = AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Test Report',
            framework=self.framework,
            requested_by=self.user,
            status='processing'
        )
        
        url = reverse('assessment-reports-generate', kwargs={'pk': report.id})
        response = self.client.post(url)
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('already being generated', response.data['error'])
    
    def test_status_check(self):
        """Test checking report generation status."""
        report = AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Test Report',
            requested_by=self.user,
            status='completed'
        )
        
        url = reverse('assessment-reports-status-check', kwargs={'pk': report.id})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['status'], 'completed')
        self.assertIn('message', response.data)

    def test_download_report_is_audited(self):
        document = Document.objects.create(
            title='Assessment report',
            uploaded_by=self.user,
            mime_type='application/pdf',
        )
        document.file.save('assessment-report.pdf', ContentFile(b'pdf bytes'), save=True)
        report = AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Completed Report',
            requested_by=self.user,
            status='completed',
            generated_file=document,
        )

        response = self.client.get(
            reverse('assessment-reports-download', kwargs={'pk': report.id}),
            HTTP_USER_AGENT='pytest-report-agent',
            REMOTE_ADDR='203.0.113.30',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(f'/api/documents/{document.id}/download/', response.data['download_url'])
        event = AuditEvent.objects.get(event='ASSESSMENT_REPORT_DOWNLOAD_REQUESTED')
        self.assertEqual(event.details['request']['ip'], '203.0.113.30')
        self.assertEqual(event.details['request']['user_agent'], 'pytest-report-agent')
        self.assertEqual(event.details['new']['filename'], 'assessment-report.pdf')
    
    @patch('exports.tasks.generate_assessment_report_task.delay')
    def test_quick_generate(self, mock_task):
        """Test creating and generating a report in one call."""
        url = reverse('assessment-reports-quick-generate')
        data = {
            'report_type': 'assessment_summary',
            'title': 'Quick Report',
            'framework': self.framework.id
        }
        
        response = self.client.post(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(AssessmentReport.objects.count(), 1)
        
        report = AssessmentReport.objects.first()
        self.assertEqual(report.status, 'processing')
        mock_task.assert_called_once_with(report.id)
    
    def test_framework_options(self):
        """Test getting framework options for reports."""
        url = reverse('assessment-reports-framework-options')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['frameworks']), 1)
        self.assertEqual(response.data['frameworks'][0]['name'], 'ISO 27001')
    
    def test_assessment_options(self):
        """Test getting assessment options for reports."""
        url = reverse('assessment-reports-assessment-options')
        response = self.client.get(url, {'framework_id': self.framework.id})
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data['assessments']), 1)
        self.assertEqual(response.data['assessments'][0]['control_id'], 'A.5.1.1')
    
    def test_compliance_gap_requires_framework(self):
        """Test that compliance gap analysis requires a framework."""
        url = reverse('assessment-reports-list')
        data = {
            'report_type': 'compliance_gap',
            'title': 'Gap Analysis',
            'description': 'Without framework'
        }
        
        response = self.client.post(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('framework', response.data)


class AssessmentReportGeneratorTest(TestCase):
    """Test AssessmentReportGenerator service."""
    
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        
        # Create test data structure
        self.framework = Framework.objects.create(
            name='Test Framework',
            short_name='TEST',
            version='1.0',
            effective_date=timezone.now().date()
        )
        
        self.clause = Clause.objects.create(
            framework=self.framework,
            clause_id='T.1',
            title='Test Clause',
            description='Test clause description'
        )
        
        self.control = Control.objects.create(
            control_id='T.1.1',
            name='Test Control',
            description='Test control description',
            control_type='preventive'
        )
        self.control.clauses.add(self.clause)
        
        self.assessment = ControlAssessment.objects.create(
            control=self.control,
            assigned_to=self.user,
            status='complete',
            implementation_status='implemented',
            implementation_approach='Test implementation'
        )
        
        # Create evidence
        self.evidence = ControlEvidence.objects.create(
            control=self.control,
            title='Test Evidence',
            evidence_type='document',
            collected_by=self.user
        )
        
        self.evidence_link = AssessmentEvidence.objects.create(
            assessment=self.assessment,
            evidence=self.evidence,
            is_primary_evidence=True
        )
    
    @patch('exports.services.HTML')
    @patch('exports.services.CSS')
    def test_generate_assessment_summary(self, mock_css, mock_html):
        """Test generating assessment summary report."""
        # Mock WeasyPrint components
        mock_html_instance = MagicMock()
        mock_html.return_value = mock_html_instance
        mock_html_instance.write_pdf.return_value = None
        
        report = AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Test Summary',
            framework=self.framework,
            requested_by=self.user
        )
        
        with patch('exports.services.AssessmentReportGenerator._save_pdf_document') as mock_save:
            mock_document = Document.objects.create(
                title='Test Report',
                uploaded_by=self.user
            )
            mock_save.return_value = mock_document
            
            generator = AssessmentReportGenerator(report)
            result = generator.generate_report()
            
            self.assertEqual(result, mock_document)
            report.refresh_from_db()
            self.assertEqual(report.status, 'completed')
            self.assertIsNotNone(report.generation_completed_at)
            event = AuditEvent.objects.get(event='ASSESSMENT_REPORT_GENERATED')
            self.assertEqual(event.details['actor']['email'], self.user.email)
            self.assertEqual(event.details['new']['status'], 'completed')
            self.assertEqual(event.details['new']['generated_file_id'], mock_document.id)
            mock_html_instance.write_pdf.assert_called_once()
            _, write_pdf_kwargs = mock_html_instance.write_pdf.call_args
            self.assertFalse(write_pdf_kwargs['presentational_hints'])
    
    def test_generate_detailed_assessment(self):
        """Test generating detailed assessment report HTML."""
        report = AssessmentReport.objects.create(
            report_type='detailed_assessment',
            title='Detailed Test',
            framework=self.framework,
            requested_by=self.user
        )
        
        generator = AssessmentReportGenerator(report)
        html_content = generator._generate_detailed_assessment()
        
        self.assertIn('Detailed Assessment Report', html_content)
        self.assertIn(self.control.control_id, html_content)
        self.assertIn(self.control.name, html_content)
        self.assertIn('Test implementation', html_content)
    
    def test_generate_evidence_portfolio(self):
        """Test generating evidence portfolio report HTML."""
        report = AssessmentReport.objects.create(
            report_type='evidence_portfolio',
            title='Evidence Portfolio',
            framework=self.framework,
            requested_by=self.user
        )
        
        generator = AssessmentReportGenerator(report)
        html_content = generator._generate_evidence_portfolio()
        
        self.assertIn('Evidence Portfolio Report', html_content)
        self.assertIn(self.evidence.title, html_content)
        self.assertIn(self.control.control_id, html_content)
    
    def test_generate_compliance_gap(self):
        """Test generating compliance gap analysis."""
        # Create some gap scenarios
        not_started = ControlAssessment.objects.create(
            control=self.control,
            assigned_to=self.user,
            status='not_started'
        )
        
        report = AssessmentReport.objects.create(
            report_type='compliance_gap',
            title='Gap Analysis',
            framework=self.framework,
            requested_by=self.user
        )
        
        generator = AssessmentReportGenerator(report)
        html_content = generator._generate_compliance_gap()
        
        self.assertIn('Compliance Gap Analysis', html_content)
        self.assertIn('Not Started', html_content)
    
    def test_generate_unknown_report_type(self):
        """Test handling unknown report type."""
        report = AssessmentReport.objects.create(
            report_type='unknown_type',
            title='Unknown Report',
            requested_by=self.user
        )
        
        generator = AssessmentReportGenerator(report)
        
        with self.assertRaises(ValueError):
            generator.generate_report()
    
    def test_filename_generation(self):
        """Test PDF filename generation."""
        report = AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Test Report',
            framework=self.framework,
            requested_by=self.user
        )
        
        generator = AssessmentReportGenerator(report)
        filename = generator._generate_filename()
        
        self.assertIn('assessment-summary', filename)
        self.assertIn('test', filename.lower())
        self.assertTrue(filename.endswith('.pdf'))
    
    def test_generation_error_handling(self):
        """Test error handling during report generation."""
        report = AssessmentReport.objects.create(
            report_type='assessment_summary',
            title='Error Test',
            framework=self.framework,
            requested_by=self.user
        )
        
        generator = AssessmentReportGenerator(report)
        
        # Mock a failure in PDF rendering
        with patch.object(generator, '_render_pdf', side_effect=Exception('PDF error')):
            with self.assertRaises(Exception):
                generator.generate_report()
            
            report.refresh_from_db()
            self.assertEqual(report.status, 'failed')
            self.assertEqual(report.error_message, 'PDF error')


class TenantDataExportAPITest(ExportTenantAPITestCase):
    """Test tenant data export API endpoints."""

    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(
            username='exporter',
            email='exporter@example.com',
            password='testpass123',
        )
        self.other_user = User.objects.create_user(
            username='other',
            email='other@example.com',
            password='testpass123',
        )
        self.client.force_authenticate(user=self.user)

    def test_coverage_manifest_documents_all_grc_modules(self):
        url = reverse('tenant-data-exports-coverage')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        modules = {entry['module'] for entry in response.data['modules']}
        self.assertEqual(
            modules,
            {
                'identity',
                'frameworks',
                'risk',
                'vendors',
                'policies',
                'training',
                'knowledge',
                'compliance_governance',
                'assets',
                'calendar',
                'vulnerabilities',
            },
        )
        for module in response.data['modules']:
            self.assertIn('xlsx', module['formats'])
            self.assertIn('csv_zip', module['formats'])
            self.assertTrue(module['sheets'])

    @patch('exports.views.generate_tenant_data_export_task.delay')
    def test_create_export_queues_async_job(self, mock_delay):
        url = reverse('tenant-data-exports-list')
        response = self.client.post(
            url,
            {
                'title': 'Full customer export',
                'export_format': 'xlsx',
                'selected_modules': ['frameworks', 'risk'],
            },
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        data_export = TenantDataExport.objects.get()
        self.assertEqual(data_export.requested_by, self.user)
        self.assertEqual(data_export.status, 'processing')
        self.assertEqual(data_export.selected_modules, ['frameworks', 'risk'])
        mock_delay.assert_called_once_with(data_export.id)
        requested_event = AuditEvent.objects.get(event='TENANT_DATA_EXPORT_REQUESTED')
        started_event = AuditEvent.objects.get(event='TENANT_DATA_EXPORT_GENERATION_STARTED')
        self.assertEqual(requested_event.details['actor']['email'], self.user.email)
        self.assertEqual(requested_event.details['object']['type'], 'exports.TenantDataExport')
        self.assertEqual(requested_event.details['new']['selected_modules'], ['frameworks', 'risk'])
        self.assertEqual(started_event.details['previous']['status'], 'pending')
        self.assertEqual(started_event.details['new']['status'], 'processing')

    def test_download_export_is_audited(self):
        data_export = TenantDataExport.objects.create(
            title='Downloadable export',
            export_format='xlsx',
            selected_modules=['assets'],
            requested_by=self.user,
            status='completed',
        )
        document = Document.objects.create(
            title='Downloadable export',
            uploaded_by=self.user,
            mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            file_size=12,
        )
        document.file.save('tenant-export.xlsx', ContentFile(b'export bytes'), save=True)
        data_export.generated_file = document
        data_export.save(update_fields=['generated_file'])

        response = self.client.get(
            reverse('tenant-data-exports-download', kwargs={'pk': data_export.id}),
            HTTP_USER_AGENT='pytest-export-agent',
            REMOTE_ADDR='203.0.113.20',
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        event = AuditEvent.objects.get(event='TENANT_DATA_EXPORT_DOWNLOAD_REQUESTED')
        self.assertEqual(event.details['request']['ip'], '203.0.113.20')
        self.assertEqual(event.details['request']['user_agent'], 'pytest-export-agent')
        self.assertEqual(event.details['new']['filename'], 'tenant-export.xlsx')
        self.assertEqual(event.details['new']['generated_file_id'], document.id)

    def test_invalid_module_is_rejected(self):
        url = reverse('tenant-data-exports-list')
        response = self.client.post(
            url,
            {
                'title': 'Bad export',
                'export_format': 'xlsx',
                'selected_modules': ['not-a-module'],
            },
            format='json',
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('selected_modules', response.data)

    def test_users_only_see_their_own_export_jobs(self):
        TenantDataExport.objects.create(
            title='Own export',
            export_format='xlsx',
            requested_by=self.user,
        )
        TenantDataExport.objects.create(
            title='Other export',
            export_format='xlsx',
            requested_by=self.other_user,
        )

        url = reverse('tenant-data-exports-list')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        titles = {row['title'] for row in response_results(response.data)}
        self.assertEqual(titles, {'Own export'})

    def test_unauthenticated_users_cannot_create_exports(self):
        self.client.force_authenticate(user=None)
        response = self.client.post(
            reverse('tenant-data-exports-list'),
            {'title': 'Anonymous export', 'export_format': 'xlsx'},
            format='json',
        )

        self.assertIn(
            response.status_code,
            [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN],
        )


class TenantDataExportGeneratorTest(TestCase):
    """Test tenant data export generation."""

    def setUp(self):
        self.user = User.objects.create_user(
            username='generator',
            email='generator@example.com',
            password='testpass123',
        )
        self.framework = Framework.objects.create(
            name='ISO 27001',
            short_name='ISO27001',
            description='Information security management',
            issuing_organization='ISO',
            version='2022',
            effective_date=timezone.now().date(),
            status='active',
            created_by=self.user,
        )
        self.asset = Asset.objects.create(
            asset_id='AST-001',
            name='Customer portal',
            asset_type='application',
            owner=self.user,
            created_by=self.user,
        )
        self.risk = Risk.objects.create(
            risk_id='RISK-001',
            title='Portal outage',
            description='Customer portal outage risk',
            impact=4,
            likelihood=3,
            created_by=self.user,
            risk_owner=self.user,
        )
        self.artefact = GovernanceArtefact.objects.create(
            artefact_id='GOV-001',
            title='ISMS scope statement',
            artefact_type='scope_document',
            owner=self.user,
            created_by=self.user,
        )
        self.requirement = RegulatoryRequirement.objects.create(
            requirement_id='REQ-001',
            title='Maintain records of processing activities',
            source_type='regulation',
            applicability_status='applicable',
            compliance_status='in_progress',
            owner=self.user,
            created_by=self.user,
        )

    def test_generate_xlsx_export_includes_selected_module_records(self):
        data_export = TenantDataExport.objects.create(
            title='Audit export',
            export_format='xlsx',
            selected_modules=['frameworks', 'risk', 'assets'],
            requested_by=self.user,
        )

        document = TenantDataExportGenerator(data_export).generate_export()
        data_export.refresh_from_db()

        self.assertEqual(data_export.status, 'completed')
        self.assertEqual(document.mime_type, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertEqual(data_export.record_counts['catalogs.Framework'], 1)
        self.assertEqual(data_export.record_counts['risk.Risk'], 1)
        self.assertEqual(data_export.record_counts['assets.Asset'], 1)
        generated_event = AuditEvent.objects.get(event='TENANT_DATA_EXPORT_GENERATED')
        self.assertEqual(generated_event.details['actor']['email'], self.user.email)
        self.assertEqual(generated_event.details['previous']['status'], 'processing')
        self.assertEqual(generated_event.details['new']['status'], 'completed')
        self.assertEqual(generated_event.details['new']['generated_file_id'], document.id)

        workbook = load_workbook(BytesIO(document.file.read()), read_only=True)
        self.assertIn('Frameworks', workbook.sheetnames)
        self.assertIn('Risks', workbook.sheetnames)
        self.assertIn('Assets', workbook.sheetnames)
        risk_rows = list(workbook['Risks'].iter_rows(values_only=True))
        self.assertIn('risk_id', risk_rows[0])
        self.assertIn('RISK-001', risk_rows[1])

    def test_generate_xlsx_export_includes_audit_records(self):
        AuditEvent.objects.create(
            user=self.user,
            event='DOCUMENT_UPLOADED',
            details={
                'event': 'DOCUMENT_UPLOADED',
                'object': {'display': 'document:1'},
            },
        )
        data_export = TenantDataExport.objects.create(
            title='Audit trail export',
            export_format='xlsx',
            selected_modules=['identity'],
            requested_by=self.user,
        )

        document = TenantDataExportGenerator(data_export).generate_export()

        workbook = load_workbook(BytesIO(document.file.read()), read_only=True)
        self.assertIn('Audit events', workbook.sheetnames)
        audit_rows = list(workbook['Audit events'].iter_rows(values_only=True))
        self.assertIn('event', audit_rows[0])
        event_index = audit_rows[0].index('event')
        self.assertIn('DOCUMENT_UPLOADED', [row[event_index] for row in audit_rows[1:]])

    def test_generate_export_failure_is_audited(self):
        data_export = TenantDataExport.objects.create(
            title='Broken export',
            export_format='xlsx',
            selected_modules=['assets'],
            requested_by=self.user,
        )
        data_export.export_format = 'bad_format'

        with self.assertRaises(ValueError):
            TenantDataExportGenerator(data_export).generate_export()

        data_export.refresh_from_db()
        self.assertEqual(data_export.status, 'failed')
        failed_event = AuditEvent.objects.get(event='TENANT_DATA_EXPORT_FAILED')
        self.assertEqual(failed_event.details['actor']['email'], self.user.email)
        self.assertEqual(failed_event.details['new']['status'], 'failed')
        self.assertIn('Unsupported export format', failed_event.details['reason'])

    def test_cleanup_old_tenant_data_exports_is_audited(self):
        data_export = TenantDataExport.objects.create(
            title='Expired export',
            export_format='xlsx',
            selected_modules=['assets'],
            requested_by=self.user,
            status='completed',
        )
        document = Document.objects.create(
            title='Expired export',
            uploaded_by=self.user,
            mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        document.file.save('expired-export.xlsx', ContentFile(b'expired'), save=True)
        data_export.generated_file = document
        data_export.save(update_fields=['generated_file'])
        TenantDataExport.objects.filter(pk=data_export.pk).update(
            requested_at=timezone.now() - timezone.timedelta(days=60)
        )

        result = cleanup_old_tenant_data_exports(days_old=30)

        self.assertEqual(result['deleted_count'], 1)
        self.assertFalse(TenantDataExport.objects.filter(pk=data_export.pk).exists())
        event = AuditEvent.objects.get(event='TENANT_DATA_EXPORT_EXPIRED')
        self.assertEqual(event.details['actor']['email'], self.user.email)
        self.assertEqual(event.details['previous']['status'], 'completed')

    def test_generate_xlsx_export_includes_governance_records(self):
        data_export = TenantDataExport.objects.create(
            title='Governance export',
            export_format='xlsx',
            selected_modules=['compliance_governance'],
            requested_by=self.user,
        )

        document = TenantDataExportGenerator(data_export).generate_export()
        data_export.refresh_from_db()

        self.assertEqual(data_export.status, 'completed')
        self.assertEqual(data_export.record_counts['compliance.GovernanceArtefact'], 1)
        self.assertEqual(data_export.record_counts['compliance.RegulatoryRequirement'], 1)

        workbook = load_workbook(BytesIO(document.file.read()), read_only=True)
        self.assertIn('Governance artefacts', workbook.sheetnames)
        self.assertIn('Regulatory requirements', workbook.sheetnames)
        requirement_rows = list(workbook['Regulatory requirements'].iter_rows(values_only=True))
        self.assertIn('requirement_id', requirement_rows[0])
        self.assertIn('REQ-001', requirement_rows[1])

    def test_serializer_download_url_uses_document_download_route(self):
        data_export = TenantDataExport.objects.create(
            title='Download route export',
            export_format='xlsx',
            selected_modules=['assets'],
            requested_by=self.user,
        )

        TenantDataExportGenerator(data_export).generate_export()
        data_export.refresh_from_db()

        from .serializers import TenantDataExportSerializer

        download_url = TenantDataExportSerializer(data_export).data['download_url']
        self.assertEqual(
            download_url,
            f'/api/documents/{data_export.generated_file_id}/download/',
        )

    def test_generate_csv_zip_export_includes_one_csv_per_sheet(self):
        data_export = TenantDataExport.objects.create(
            title='CSV export',
            export_format='csv_zip',
            selected_modules=['assets'],
            requested_by=self.user,
        )

        document = TenantDataExportGenerator(data_export).generate_export()
        data_export.refresh_from_db()

        self.assertEqual(data_export.status, 'completed')
        self.assertEqual(document.mime_type, 'application/zip')
        with zipfile.ZipFile(BytesIO(document.file.read())) as archive:
            self.assertIn('Assets.csv', archive.namelist())
            assets_csv = archive.read('Assets.csv').decode('utf-8')
            self.assertIn('asset_id', assets_csv)
            self.assertIn('AST-001', assets_csv)

    def test_export_manifest_matches_documented_service_coverage(self):
        modules = {entry['module'] for entry in get_export_coverage_manifest()}

        self.assertIn('frameworks', modules)
        self.assertIn('vulnerabilities', modules)
        self.assertIn('calendar', modules)
