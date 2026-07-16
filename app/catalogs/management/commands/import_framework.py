import json
import yaml
import hashlib
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.utils.dateparse import parse_date
from django.db import transaction
from openpyxl import load_workbook
from core.audit import log_audit_event
from catalogs.models import Framework, Clause, Control
from django.contrib.auth import get_user_model

User = get_user_model()


class Command(BaseCommand):
    help = 'Import compliance frameworks from structured JSON, YAML, or XLSX files'
    
    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            help='Path to framework definition file (JSON, YAML, or XLSX)'
        )
        parser.add_argument(
            '--update',
            action='store_true',
            help='Update existing framework if it already exists'
        )
        parser.add_argument(
            '--user',
            type=str,
            help='Username of user importing the framework',
            default='system'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Show what would be imported without actually importing'
        )
        parser.add_argument(
            '--name',
            type=str,
            help='Framework name for spreadsheet imports'
        )
        parser.add_argument(
            '--short-name',
            type=str,
            help='Framework short name for spreadsheet imports'
        )
        parser.add_argument(
            '--framework-version',
            dest='framework_version',
            type=str,
            help='Framework version for spreadsheet imports'
        )
        parser.add_argument(
            '--issuing-organization',
            type=str,
            help='Issuing organisation for spreadsheet imports'
        )
        parser.add_argument(
            '--effective-date',
            type=str,
            help='Effective date for spreadsheet imports in YYYY-MM-DD format'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            help='Worksheet name for spreadsheet imports. Defaults to the first non-empty sheet.'
        )
    
    def handle(self, *args, **options):
        file_path = Path(options['file_path'])
        
        if not file_path.exists():
            raise CommandError(f'File not found: {file_path}')
        
        # Determine file format
        if file_path.suffix.lower() == '.json':
            framework_data = self.load_json_file(file_path)
        elif file_path.suffix.lower() in ['.yaml', '.yml']:
            framework_data = self.load_yaml_file(file_path)
        elif file_path.suffix.lower() == '.xlsx':
            framework_data = self.load_xlsx_file(file_path, options)
        else:
            raise CommandError('Unsupported file format. Use .json, .yaml, .yml, or .xlsx files.')
        
        # Validate framework data
        self.validate_framework_data(framework_data)
        
        # Get user
        try:
            if options['user'] == 'system':
                user = None
            else:
                user = User.objects.get(username=options['user'])
        except User.DoesNotExist:
            raise CommandError(f'User not found: {options["user"]}')
        
        # Calculate file checksum for tracking changes
        file_checksum = self.calculate_file_checksum(file_path)
        
        if options['dry_run']:
            self.show_dry_run_summary(framework_data, file_checksum)
            return
        
        # Import framework
        try:
            with transaction.atomic():
                framework = self.import_framework_data(
                    framework_data, 
                    file_path, 
                    file_checksum, 
                    user, 
                    options['update']
                )
                self.record_import_audit_event(
                    framework=framework,
                    framework_data=framework_data,
                    file_path=file_path,
                    file_checksum=file_checksum,
                    user=user,
                    update_existing=options['update'],
                )
                
                self.stdout.write(
                    self.style.SUCCESS(
                        f'Successfully imported framework: {framework.name} v{framework.version}'
                    )
                )
                self.stdout.write(f'Framework ID: {framework.id}')
                self.stdout.write(f'Total clauses imported: {framework.clause_count}')
                
        except Exception as e:
            raise CommandError(f'Failed to import framework: {str(e)}')
    
    def load_json_file(self, file_path):
        """Load framework data from JSON file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            raise CommandError(f'Invalid JSON file: {e}')
        except Exception as e:
            raise CommandError(f'Error reading file: {e}')
    
    def load_yaml_file(self, file_path):
        """Load framework data from YAML file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except yaml.YAMLError as e:
            raise CommandError(f'Invalid YAML file: {e}')
        except Exception as e:
            raise CommandError(f'Error reading file: {e}')

    def load_xlsx_file(self, file_path, options):
        """Load framework data from a spreadsheet with common compliance headers."""
        metadata = self.build_spreadsheet_metadata(file_path, options)
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = self.get_spreadsheet_worksheet(workbook, options.get('sheet'))
        header_row, header_map = self.find_header_row(worksheet)

        clauses = []
        controls = []
        seen_clause_ids = set()

        for sort_order, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=10,
        ):
            row_data = self.extract_spreadsheet_row(row, header_map)
            if not row_data:
                continue

            clause_id = row_data['clause_id']
            title = row_data['title']
            description = row_data['description']
            testing_procedures = row_data.get('testing_procedures', '')
            guidance = row_data.get('implementation_guidance', '')

            if clause_id not in seen_clause_ids:
                clauses.append({
                    'clause_id': clause_id,
                    'title': title,
                    'description': description,
                    'clause_type': row_data.get('clause_type', 'control'),
                    'criticality': row_data.get('criticality', 'medium'),
                    'sort_order': sort_order,
                    'parent_clause_id': row_data.get('parent_clause_id') or self.infer_parent_clause_id(clause_id),
                    'implementation_guidance': guidance,
                    'testing_procedures': testing_procedures,
                    'external_references': row_data.get('external_references', {}),
                })
                seen_clause_ids.add(clause_id)

            if self.should_create_control(row_data):
                controls.append({
                    'control_id': row_data.get('control_id') or self.build_control_id(metadata['short_name'], clause_id),
                    'name': row_data.get('control_name') or title,
                    'description': description,
                    'control_type': row_data.get('control_type', 'administrative'),
                    'status': row_data.get('control_status', 'planned'),
                    'clauses': [clause_id],
                    'implementation_details': guidance,
                    'evidence_requirements': row_data.get('evidence_requirements') or testing_procedures,
                    'documentation_links': row_data.get('documentation_links', []),
                    'risk_rating': row_data.get('risk_rating', ''),
                })

        if not clauses:
            raise CommandError(f'No importable clauses found in worksheet: {worksheet.title}')

        return {
            **metadata,
            'description': metadata.get('description') or f'Imported from spreadsheet {file_path.name}',
            'framework_type': 'financial' if 'pci' in metadata['short_name'].lower() else 'security',
            'status': 'active',
            'clauses': clauses,
            'controls': controls,
        }

    def build_spreadsheet_metadata(self, file_path, options):
        """Build framework metadata required by the common import pipeline."""
        missing = [
            option_name
            for option_name, cli_name in [
                ('name', '--name'),
                ('framework_version', '--framework-version'),
                ('issuing_organization', '--issuing-organization'),
                ('effective_date', '--effective-date'),
            ]
            if not options.get(option_name)
        ]
        if missing:
            cli_options = ', '.join(f'--{name.replace("_", "-")}' for name in missing)
            raise CommandError(
                f'Spreadsheet imports require framework metadata: {cli_options}'
            )

        return {
            'name': options['name'],
            'short_name': options.get('short_name') or self.slugify_identifier(file_path.stem)[:50],
            'version': options['framework_version'],
            'issuing_organization': options['issuing_organization'],
            'effective_date': options['effective_date'],
            'external_id': self.slugify_identifier(file_path.stem),
            'imported_from': str(file_path),
        }

    def get_spreadsheet_worksheet(self, workbook, sheet_name):
        """Return the requested worksheet, or the first worksheet with data."""
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f'Worksheet not found: {sheet_name}')
            return workbook[sheet_name]

        for worksheet in workbook.worksheets:
            if worksheet.max_row > 1 and worksheet.max_column > 1:
                return worksheet
        raise CommandError('No non-empty worksheet found in spreadsheet.')

    def find_header_row(self, worksheet):
        """Find a row containing a recognised clause/control identifier header."""
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            header_map = {}
            for index, value in enumerate(row):
                normalized = self.normalize_header(value)
                canonical = self.canonical_header(normalized)
                if canonical:
                    header_map[canonical] = index

            if 'clause_id' in header_map and ('description' in header_map or 'title' in header_map):
                return row_number, header_map

        raise CommandError(
            'Could not find spreadsheet header row. Expected a recognised ID column and a requirement/title column.'
        )

    def extract_spreadsheet_row(self, row, header_map):
        """Normalize a spreadsheet row into clause/control import fields."""
        raw_clause_id = self.get_cell(row, header_map, 'clause_id')
        description = self.get_cell(row, header_map, 'description') or self.get_cell(row, header_map, 'title')
        if not raw_clause_id or not description:
            return {}

        clause_id = self.normalize_identifier(raw_clause_id)
        description = self.clean_text(description)
        title = self.clean_text(self.get_cell(row, header_map, 'title')) or self.extract_title(description, clause_id)

        return {
            'clause_id': clause_id,
            'title': title,
            'description': description,
            'parent_clause_id': self.normalize_optional_identifier(self.get_cell(row, header_map, 'parent_clause_id')),
            'testing_procedures': self.clean_text(self.get_cell(row, header_map, 'testing_procedures')),
            'implementation_guidance': self.clean_text(self.get_cell(row, header_map, 'implementation_guidance')),
            'evidence_requirements': self.clean_text(self.get_cell(row, header_map, 'evidence_requirements')),
            'criticality': self.normalize_choice(self.get_cell(row, header_map, 'criticality'), {'low', 'medium', 'high', 'critical'}, 'medium'),
            'clause_type': self.normalize_choice(
                self.get_cell(row, header_map, 'clause_type'),
                {'control', 'policy', 'procedure', 'documentation', 'assessment', 'monitoring', 'reporting'},
                'control',
            ),
            'control_id': self.normalize_optional_identifier(self.get_cell(row, header_map, 'control_id')),
            'control_name': self.clean_text(self.get_cell(row, header_map, 'control_name')),
            'control_type': self.normalize_choice(
                self.get_cell(row, header_map, 'control_type'),
                {'preventive', 'detective', 'corrective', 'compensating', 'administrative', 'technical', 'physical'},
                'administrative',
            ),
            'control_status': self.normalize_choice(
                self.get_cell(row, header_map, 'control_status'),
                {'planned', 'in_progress', 'implemented', 'testing', 'active', 'remediation', 'disabled', 'retired'},
                'planned',
            ),
            'risk_rating': self.normalize_choice(
                self.get_cell(row, header_map, 'risk_rating'),
                {'low', 'medium', 'high', 'critical'},
                '',
            ),
        }

    def canonical_header(self, normalized_header):
        """Map common spreadsheet headers to importer field names."""
        header_aliases = {
            'pci dss id': 'clause_id',
            'id': 'clause_id',
            'requirement id': 'clause_id',
            'clause id': 'clause_id',
            'control id': 'control_id',
            'parent id': 'parent_clause_id',
            'parent clause id': 'parent_clause_id',
            'defined approach requirements': 'description',
            'requirement': 'description',
            'requirements': 'description',
            'description': 'description',
            'control objective': 'description',
            'title': 'title',
            'control title': 'control_name',
            'name': 'title',
            'defined approach testing procedures': 'testing_procedures',
            'testing procedures': 'testing_procedures',
            'test procedures': 'testing_procedures',
            'evidence': 'evidence_requirements',
            'evidence requirements': 'evidence_requirements',
            'evidence required': 'evidence_requirements',
            'guidance': 'implementation_guidance',
            'implementation guidance': 'implementation_guidance',
            'criticality': 'criticality',
            'risk rating': 'risk_rating',
            'clause type': 'clause_type',
            'control type': 'control_type',
            'control status': 'control_status',
        }
        return header_aliases.get(normalized_header)

    def normalize_header(self, value):
        return re.sub(r'\s+', ' ', str(value or '').strip().lower())

    def get_cell(self, row, header_map, field_name):
        index = header_map.get(field_name)
        if index is None or index >= len(row):
            return ''
        return row[index]

    def clean_text(self, value):
        if value is None:
            return ''
        return re.sub(r'\s+', ' ', str(value).strip())

    def normalize_identifier(self, value):
        if value is None or value == '':
            return ''
        if isinstance(value, float | int):
            value = Decimal(str(value))
        if isinstance(value, Decimal):
            return format(value.normalize(), 'f')
        return str(value).strip()

    def normalize_optional_identifier(self, value):
        normalized = self.normalize_identifier(value)
        return normalized or ''

    def slugify_identifier(self, value):
        return re.sub(r'[^A-Z0-9]+', '-', str(value).upper()).strip('-')

    def build_control_id(self, framework_short_name, clause_id):
        return self.slugify_identifier(f'{framework_short_name}-{clause_id}')[:50]

    def infer_parent_clause_id(self, clause_id):
        if '.' not in clause_id:
            return ''
        return clause_id.rsplit('.', 1)[0]

    def extract_title(self, description, clause_id):
        text = re.sub(rf'^{re.escape(clause_id)}\s*', '', description).strip()
        first_sentence = re.split(r'(?<=[.!?])\s+', text, maxsplit=1)[0]
        return first_sentence[:300] or clause_id

    def normalize_choice(self, value, valid_choices, default):
        normalized = self.clean_text(value).lower().replace('-', '_').replace(' ', '_')
        return normalized if normalized in valid_choices else default

    def should_create_control(self, row_data):
        return bool(
            row_data.get('control_id')
            or row_data.get('testing_procedures')
            or row_data.get('evidence_requirements')
        )
    
    def calculate_file_checksum(self, file_path):
        """Calculate SHA256 checksum of the file."""
        sha256_hash = hashlib.sha256()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)
        return sha256_hash.hexdigest()
    
    def validate_framework_data(self, data):
        """Validate that framework data has required fields."""
        required_fields = ['name', 'version', 'issuing_organization', 'effective_date']
        
        for field in required_fields:
            if field not in data:
                raise CommandError(f'Missing required field: {field}')
        
        # Validate clauses structure
        if 'clauses' in data and not isinstance(data['clauses'], list):
            raise CommandError('Clauses must be a list')
        
        # Validate each clause
        for i, clause in enumerate(data.get('clauses', [])):
            if not isinstance(clause, dict):
                raise CommandError(f'Clause {i} must be a dictionary')
            
            clause_required = ['clause_id', 'title', 'description']
            for field in clause_required:
                if field not in clause:
                    raise CommandError(f'Clause {i} missing required field: {field}')
    
    def show_dry_run_summary(self, framework_data, checksum):
        """Show what would be imported without actually importing."""
        self.stdout.write(self.style.WARNING('DRY RUN - No changes will be made'))
        self.stdout.write('-' * 50)
        
        self.stdout.write(f'Framework Name: {framework_data["name"]}')
        self.stdout.write(f'Version: {framework_data["version"]}')
        self.stdout.write(f'Issuing Organisation: {framework_data["issuing_organization"]}')
        self.stdout.write(f'Effective Date: {framework_data["effective_date"]}')
        self.stdout.write(f'File Checksum: {checksum[:16]}...')
        
        clauses = framework_data.get('clauses', [])
        controls = framework_data.get('controls', [])
        self.stdout.write(f'Total Clauses: {len(clauses)}')
        self.stdout.write(f'Total Controls: {len(controls)}')
        
        if clauses:
            self.stdout.write('\nSample clauses:')
            for clause in clauses[:3]:  # Show first 3 clauses
                self.stdout.write(f'  - {clause["clause_id"]}: {clause["title"]}')
            
            if len(clauses) > 3:
                self.stdout.write(f'  ... and {len(clauses) - 3} more clauses')
    
    def import_framework_data(self, data, file_path, checksum, user, update_existing):
        """Import framework and clauses from validated data."""
        
        # Check if framework already exists
        existing_framework = Framework.objects.filter(
            name=data['name'], 
            version=data['version']
        ).first()
        
        if existing_framework and not update_existing:
            raise CommandError(
                f'Framework {data["name"]} v{data["version"]} already exists. '
                'Use --update to update existing framework.'
            )
        
        # Parse effective date
        effective_date = parse_date(data['effective_date'])
        if not effective_date:
            raise CommandError(f'Invalid effective_date format: {data["effective_date"]}')
        
        # Parse expiry date if provided
        expiry_date = None
        if 'expiry_date' in data and data['expiry_date']:
            expiry_date = parse_date(data['expiry_date'])
            if not expiry_date:
                raise CommandError(f'Invalid expiry_date format: {data["expiry_date"]}')
        
        # Create or update framework
        framework_defaults = {
            'short_name': data.get('short_name', data['name'][:50]),
            'description': data.get('description', ''),
            'framework_type': data.get('framework_type', 'security'),
            'external_id': data.get('external_id', ''),
            'issuing_organization': data['issuing_organization'],
            'official_url': data.get('official_url', ''),
            'effective_date': effective_date,
            'expiry_date': expiry_date,
            'status': data.get('status', 'active'),
            'is_mandatory': data.get('is_mandatory', False),
            'created_by': user,
            'imported_from': str(file_path),
            'import_checksum': checksum,
        }
        
        if existing_framework and update_existing:
            # Update existing framework
            for field, value in framework_defaults.items():
                if field != 'created_by':  # Don't change creator
                    setattr(existing_framework, field, value)
            existing_framework.save()
            framework = existing_framework
            
            # Clear existing clauses for update
            framework.clauses.all().delete()
            self.stdout.write('Updated existing framework')
        else:
            # Create new framework
            framework = Framework.objects.create(
                name=data['name'],
                version=data['version'],
                **framework_defaults
            )
            self.stdout.write('Created new framework')
        
        # Import clauses
        clauses_data = data.get('clauses', [])
        self.import_clauses(framework, clauses_data)

        # Import controls after clauses so many-to-many links are stable.
        controls_data = data.get('controls', [])
        self.import_controls(framework, controls_data, user)
        
        return framework

    def record_import_audit_event(
        self,
        *,
        framework,
        framework_data,
        file_path,
        file_checksum,
        user,
        update_existing,
    ):
        """Record the catalogue import in the tenant audit trail."""
        log_audit_event(
            actor=user,
            event='FRAMEWORK_IMPORTED',
            target=framework,
            object_display=f'{framework.name} {framework.version}',
            source={'type': 'import', 'reference': str(file_path)},
            details={
                'framework_id': framework.id,
                'framework_name': framework.name,
                'framework_version': framework.version,
                'source_file': str(file_path),
                'source_checksum': file_checksum,
                'updated_existing': update_existing,
                'clause_count': len(framework_data.get('clauses', [])),
                'control_count': len(framework_data.get('controls', [])),
            },
        )
    
    def import_clauses(self, framework, clauses_data):
        """Import clauses for the framework."""
        clause_objects = []
        clause_map = {}  # For handling parent-child relationships
        
        # Sort clauses by sort_order or clause_id for consistent processing
        sorted_clauses = sorted(
            clauses_data, 
            key=lambda x: (x.get('sort_order', 999), x['clause_id'])
        )
        
        # First pass: create all clauses without parent relationships
        for clause_data in sorted_clauses:
            clause = Clause(
                framework=framework,
                clause_id=clause_data['clause_id'],
                title=clause_data['title'],
                description=clause_data['description'],
                clause_type=clause_data.get('clause_type', 'control'),
                criticality=clause_data.get('criticality', 'medium'),
                is_testable=clause_data.get('is_testable', True),
                sort_order=clause_data.get('sort_order', 0),
                implementation_guidance=clause_data.get('implementation_guidance', ''),
                testing_procedures=clause_data.get('testing_procedures', ''),
                external_references=clause_data.get('external_references', {}),
            )
            clause_objects.append(clause)
            clause_map[clause_data['clause_id']] = clause
        
        # Bulk create clauses
        Clause.objects.bulk_create(clause_objects)
        
        # Second pass: update parent relationships
        clauses_with_parents = [
            c for c in sorted_clauses 
            if c.get('parent_clause_id')
        ]
        
        if clauses_with_parents:
            # Reload clauses from database to get IDs
            db_clauses = {
                c.clause_id: c for c in 
                Clause.objects.filter(framework=framework)
            }
            
            for clause_data in clauses_with_parents:
                parent_clause_id = clause_data['parent_clause_id']
                if parent_clause_id in db_clauses:
                    clause = db_clauses[clause_data['clause_id']]
                    clause.parent_clause = db_clauses[parent_clause_id]
                    clause.save()
                else:
                    self.stdout.write(
                        self.style.WARNING(
                            f'Parent clause {parent_clause_id} not found for {clause_data["clause_id"]}'
                        )
                    )
        
        self.stdout.write(f'Imported {len(clause_objects)} clauses')

    def import_controls(self, framework, controls_data, user):
        """Import controls and link them to already imported clauses."""
        if not controls_data:
            self.stdout.write('Imported 0 controls')
            return

        db_clauses = {
            clause.clause_id: clause
            for clause in Clause.objects.filter(framework=framework)
        }
        imported_count = 0

        for control_data in controls_data:
            clause_ids = control_data.get('clauses', [])
            linked_clauses = [
                db_clauses[clause_id]
                for clause_id in clause_ids
                if clause_id in db_clauses
            ]
            if not linked_clauses:
                self.stdout.write(
                    self.style.WARNING(
                        f'Skipping control {control_data["control_id"]}: no matching clauses found'
                    )
                )
                continue

            control_defaults = {
                'name': control_data['name'][:200],
                'description': control_data.get('description', ''),
                'control_type': control_data.get('control_type', 'administrative'),
                'status': control_data.get('status', 'planned'),
                'implementation_details': control_data.get('implementation_details', ''),
                'evidence_requirements': control_data.get('evidence_requirements', ''),
                'documentation_links': control_data.get('documentation_links', []),
                'risk_rating': control_data.get('risk_rating', ''),
                'created_by': user,
            }
            control, _ = Control.objects.update_or_create(
                control_id=control_data['control_id'],
                defaults=control_defaults,
            )
            control.clauses.add(*linked_clauses)
            imported_count += 1

        self.stdout.write(f'Imported {imported_count} controls')
